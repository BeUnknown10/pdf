# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "output.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj['A1': 'B5']

# Print value of cell object
# using the value attribute
for cell1, cell2 in cell_obj:
	print(cell1.value, cell2.value)

from openpyxl import Workbook

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
workbook = Workbook()

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
workbook.save(filename="sample.xlsx")

# import openpyxl module


# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute
sheet = wb.active

# Cell objects also have row, column
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or column integer
# is 1, not 0. Cell object is created by
# using sheet object's cell() method.
c1 = sheet.cell(row = 1, column = 1)

# writing values to cells
c1.value = "State"

c2 = sheet.cell(row= 1 , column = 2)
c2.value = "code"

# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
c3 = sheet['A2']
c3.value = "Karnataka"

# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "Ka"

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wb.save("sample.xlsx")

# import openpyxl module
import openpyxl

wb = openpyxl.load_workbook("sample.xlsx")

sheet = wb.active

c = sheet['A3']
c.value = "Maharasthra"
d = sheet['B3']
d.value = "MH"
e = sheet['A4']
e.value = "Tamilnadu"
f = sheet['B4']
f.value = "TN"
wb.save("sample.xlsx")
